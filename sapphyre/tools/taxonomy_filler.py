import csv
from math import ceil
import openpyxl
import requests
import json
import argparse
from bs4 import BeautifulSoup
from multiprocessing.pool import ThreadPool as Pool
from ..utils import printv

def main(args):
    input = args.INPUT
    processes = args.processes
    output_fails = args.output_fails
    verbose = args.verbose
    column_name = args.organism_col

    out = "Filled_" + input if not args.output else args.output

    output_type = out.split('.')[-1]

    rows = []
    to_scrape = []


    wb = None
    if input.split('.')[-1] == "csv":
        input_type = "csv"
        with open(input, "r") as f:
            csv_reader = csv.reader(f, delimiter=',', quotechar='"')
            for i,row in enumerate(csv_reader):
                if i == 0:
                    order = []
                    for val in row:
                        order.append(val)
                    
                    if column_name not in order:
                        target_col = list(map(lambda x: str(x).lower(), order)).index(column_name.lower())
                        if target_col == -1:
                            printv(f"Column {column_name} not found in the input file. Use -col to set organism col", verbose)
                            return False
                    else:
                        target_col = order.index(column_name)
                else:
                    rows.append(row)
                    organism_name = row[target_col]
                    to_scrape.append(organism_name)
    else:
        input_type = "xlsx"
        wb = openpyxl.load_workbook(input)
        sheet = wb[wb.sheetnames[0]]
        for i,row in enumerate(sheet.rows):
            if i == 0:
                order = []
                for col in row:
                    order.append(col.value)

                if column_name not in order:
                    target_col = list(map(lambda x: str(x).lower(), order)).index(column_name.lower())
                    if target_col == -1:
                        printv(f"Column {column_name} not found in the input file. Use -col to set organism col", verbose)
                        return False
                else:
                    target_col = order.index(column_name)
            else:
                rows.append(row)
                organism_name = row[target_col].value.replace(" sp.","").replace(" spp.","")
                to_scrape.append(organism_name)

    to_scrape = list(set(to_scrape))

    x = ceil(len(to_scrape) / processes)
    arguments = [(to_scrape[i: i+x], verbose) for i in range(0, len(to_scrape), x)]

    if processes == 1:
        results = [scrape(i) for i in arguments]
    else:
        with Pool(processes) as p:
            results = p.starmap(scrape, arguments)

    final_dict = {}
    all_fails = set()
    for result, fails in results:
        all_fails.update(fails)
        final_dict.update(json.loads(result))

    if all_fails:
        if not output_fails:
            printv(f"Failed to scrape {len(all_fails)} organisms, use -of to output them", verbose)
        else:
            printv(f"Wrote {len(all_fails)} failed organisms", verbose)
            with open(output_fails, "w") as f:
                for fail in all_fails:
                    f.write(fail + "\n")        

    for row in rows:
        if input_type == "csv":
            if len(order) != len(row):
                row += [""] * (len(order) - len(row)
                )
            organism_name = row[target_col]
        else:
            organism_name = row[target_col].value

        this_out = final_dict.get(organism_name, {})

        for x,col in enumerate(order):
            col = str(col).lower()
            if col in this_out:
                val = this_out[col]
                if input_type == "csv":
                    if x >= len(row):
                        row.append(val)
                        continue

                    row[x] = val
                    continue

                row[x].value = val

    if output_type == "csv":
        with open(out, "w") as f:
            csv_writer = csv.writer(f, delimiter=',', quotechar='"')
            for row in rows:
                csv_writer.writerow(row)
    else:
        if input_type == "csv" and output_type == "xlsx":
            wb = openpyxl.Workbook()
            sheet = wb.active

            for i, col in enumerate(order):
                sheet.cell(row=1, column=i+1, value=col)

            for i, row in enumerate(rows):
                for j, val in enumerate(row):
                    sheet.cell(row=i+2, column=j+1, value=val)
        wb.save(out)

    return True

def scrape(organisms, verbose):
    printv(f"Thread started: Scraping {len(organisms)}", verbose)
    this_values = {}
    fails = set()
    for organism in organisms:
        if organism in this_values:
            continue
        try:
            url = 'https://www.ncbi.nlm.nih.gov/Taxonomy/Browser/wwwtax.cgi?name={}'.format(organism.split("_")[0])

            req = requests.get(url)
            printv(req.url, verbose, 2)
            soup = BeautifulSoup(req.content, "html.parser")

            for strong in soup.find_all("strong"):
                strong_content = strong.contents[0]
                if (" <subgenus>" in strong_content or 
                    "<flies>" in strong_content or 
                    "<flies,subgenus>" in strong_content or
                    "bees>" in strong_content or 
                    "bees,genus>" in strong_content or 
                    "<butterflies>" in strong_content or
                    "<beetle" in strong_content or
                    "<ants>" in strong_content or 
                    ("thrips" in organism.lower() and "<Thrips>" in strong_content)
                    ):

                    req = requests.get(f"https://www.ncbi.nlm.nih.gov/Taxonomy/Browser/{strong.parent['href']}")
                    soup = BeautifulSoup(req.content, "html.parser")

            this_out = {}
            done = []

            for a in soup.find_all("a"):
                if "wwwtax.cgi?mode=Undef" in a["href"]:
                    if a.get("title", ""):
                        field = a.get("title", "").title()
                        done.append(field)
                        if field in this_out:
                            field = f"{field}_{done.count(field)}"
                        this_out[field.lower()] = a.contents[0]

            this_values[organism] = this_out
        except:
            fails.add(organism)
    return json.dumps(this_values), fails