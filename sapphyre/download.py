import csv
import os
import sys
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from subprocess import PIPE, Popen

import openpyxl
import requests
from bs4 import BeautifulSoup
from shutil import rmtree

from .utils import printv


def download_parallel_srr(arguments):
    path_to_bin, srr_acession, path_to_download, verbose, fast, keep_prefetch, skip_prefetch = arguments
    printv(f"Download {srr_acession} to {path_to_download}...", verbose)

    command = "fasterq-dump" if fast else "fastq-dump --gzip"

    if skip_prefetch:
        path_of_download = srr_acession
    else:
        with Popen(
            f"{path_to_bin}prefetch {srr_acession} -O {path_to_download}",
            shell=True,
            stdout=PIPE,
        ) as p:
            try:
                print(p.stdout.read().decode())
            except UnicodeDecodeError:
                print(
                    "ErrUnicode decoding error, UTF-8 charset does not contain the bytecode for gotten character",
                )
                sys.exit(1)
        path_of_download = os.path.join(path_to_download, f"{srr_acession}")
    with Popen(
        f"{path_to_bin}{command} {path_of_download} --split-3 -O {path_to_download}",
        shell=True,
        stdout=PIPE,
    ) as p:
        try:
            print(p.stdout.read().decode())
        except UnicodeDecodeError:
            print(
                "ErrUnicode decoding error, UTF-8 charset does not contain the bytecode for gotten character",
            )
            sys.exit(1)
    
    if not keep_prefetch:
        rmtree(path_of_download)


def download_parallel_wgs(arguments):
    download_link, path_to_download, verbose = arguments

    file_name = download_link.split("/")[-1]

    printv(f"Download {file_name} to {path_to_download}...", verbose)
    download = requests.get(download_link)

    path = os.path.join(path_to_download, file_name)

    open(path, "wb").write(download.content)


def main(args):
    path_to_bin = ""
    if args.bin:
        path_to_bin = args.bin
    csvfile = Path(args.INPUT)
    this_suffix = csvfile.suffix

    path_to_download = Path(
        os.getcwd(), "input", csvfile.name.removesuffix(this_suffix)
    )
    os.makedirs(path_to_download, exist_ok=True)
    if args.wgs:
        if "csv" in csvfile.suffix:
            with open(csvfile, encoding="utf-8") as fp:
                csv_read = csv.reader(fp, delimiter=",", quotechar='"')
                arguments = []
                for i, fields in enumerate(csv_read):
                    if i == 0:
                        continue
                    prefix = fields[0]
                    url = f"https://www.ncbi.nlm.nih.gov/Traces/wgs/{prefix}"
                    req = requests.get(url)

                    soup = BeautifulSoup(req.content, "html.parser")
                    em = soup.find("em", text="FASTA:")
                    if not em:
                        print(f"Failed to find FASTA: {prefix}")
                        continue

                    container = em.find_parent()
                    for a in container.find_all("a", href=True):
                        if "sra-download.ncbi.nlm.nih.gov" in a["href"]:
                            print(f"Attempting to download: {a.contents[0]}")

                            arguments.append(
                                (a["href"], path_to_download, args.verbose)
                            )
        elif "xls" in csvfile.suffix:
            workbook = openpyxl.load_workbook(csvfile)
            sheet = workbook.active
            arguments = []

            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    continue
                fields = list(row)
                prefix = fields[0]
                url = f"https://www.ncbi.nlm.nih.gov/Traces/wgs/{prefix}"
                req = requests.get(url)

                soup = BeautifulSoup(req.content, "html.parser")
                em = soup.find("em", text="FASTA:")
                if not em:
                    print(f"Failed to find FASTA: {prefix}")
                    continue

                container = em.find_parent()
                for a in container.find_all("a", href=True):
                    if "sra-download.ncbi.nlm.nih.gov" in a["href"]:
                        print(f"Attempting to download: {a.contents[0]}")

                        arguments.append((a["href"], path_to_download, args.verbose))
    elif this_suffix == ".csv":
        with open(csvfile, encoding="utf-8") as fp:
            csv_read = csv.reader(fp, delimiter=",", quotechar='"')
            arguments = []
            for i, fields in enumerate(csv_read):
                out_fields = ['"{}"'.format(str(i).replace("ï»¿", "")) for i in fields]
                if fields[0] == "Experiment Accession":
                    out_fields.append('"SRR Acession"')

                elif fields[0] != "":
                    acession = fields[0]
                    print(f"Searching for runs in SRA: {acession}")

                    url = f"https://www.ncbi.nlm.nih.gov/sra/{acession}[accn]"
                    # TODO handle network errors
                    req = requests.get(url, timeout=500)
                    soup = BeautifulSoup(req.content, "html.parser")
                    for srr_acession in (
                        a.contents[0]
                        for a in soup.find_all("a", href=True)
                        if a["href"].startswith("//trace.ncbi.nlm.nih.gov/Traces?run")
                    ):
                        print(f"Attempting to download: {srr_acession}")
                        out_fields.append(f'"{srr_acession}"')

                        # TODO: verify download is successful
                        arguments.append(
                            (path_to_bin, srr_acession, path_to_download, args.verbose, args.fast, args.keep_prefetch, args.skip_prefetch,),
                        )
    elif "xls" in this_suffix:
        workbook = openpyxl.load_workbook(csvfile)
        sheet = workbook.active
        arguments = []

        for row in sheet.iter_rows(values_only=True):
            fields = list(row)
            out_fields = ['"{}"'.format(str(i).replace("ï»¿", "")) for i in fields]

            if fields[0] == "Experiment Accession":
                out_fields.append('"SRR Accession"')
            elif fields[0] != "":
                accession = fields[0]
                print(f"Searching for runs in SRA: {accession}")
                url = f"https://www.ncbi.nlm.nih.gov/sra/{accession}[accn]"

                # TODO handle network errors
                req = requests.get(url, timeout=500)
                soup = BeautifulSoup(req.content, "html.parser")

                for srr_accession in (
                    a.contents[0]
                    for a in soup.find_all("a", href=True)
                    if a["href"].startswith("//trace.ncbi.nlm.nih.gov/Traces?run")
                ):
                    print(f"Attempting to download: {srr_accession}")
                    out_fields.append(f'"{srr_accession}"')

                    # TODO: verify download is successful
                    arguments.append(
                        (path_to_bin, srr_accession, path_to_download, args.verbose, args.fast, args.keep_prefetch, args.skip_prefetch,)
                    )
    elif not this_suffix:
        arguments = [
            (path_to_bin, args.INPUT, path_to_download, args.verbose, args.fast, args.keep_prefetch, args.skip_prefetch,)
        ]
    func = download_parallel_srr if not args.wgs else download_parallel_wgs
    with ThreadPoolExecutor(args.processes) as pool:
        pool.map(func, arguments, chunksize=1)

    return True
